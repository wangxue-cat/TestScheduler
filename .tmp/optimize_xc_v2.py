import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

hfont = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
cfont = Font(name='微软雅黑', size=10)
wrap = Alignment(wrap_text=True, vertical='top')
cwrap = Alignment(horizontal='center', vertical='top', wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
hdrs = ['STORY编号','节点路径','用例名称','摘要','前置条件','执行步骤','预期结果','实际结果','优先级','作者','类型','是否需要回归','应用','标签']

N = '/全部需求/产品需求/API/携程/协议'
cases = []

def c(story, name, summary, precond, steps, expected, pri, reg, tags, author='王雪', app='aps-app'):
    cases.append([story, N, name, summary, precond, steps, expected, '', pri, author, '功能测试', reg, app, tags])

# ===== DB config reference =====
# partner_agreement_config: enabled=Y/N, grayscale_ratio(>1000=off), mobile_whitelist, id_number_whitelist, order_no_whitelist, timeout_minutes
# partner_agreement_instance: status=INIT/SUCCESS/FAIL, file_id, ags_status, retry_count
# New mode: enabled=Y → 拉取携程侧协议； Old mode: enabled=N + grayscale>1000 → 走原有ags协议

# ============ A. 新模式 - 授信进件 新户 (8条) ============
c('JYSG-149007',
  '携程新模式-授信进件提交-新户-协议拉取成功',
  '验证partner_agreement_config.enabled=Y下，新户授信进件提交时成功拉取携程侧用户章协议，检查lps日志无ags协议调用',
  '1. aps_stg3.partner_agreement_config表：enabled=Y, grayscale_ratio>1000, partner_code=XCtrip\n2. 携程渠道新用户（未在partner_agreement_mode_record中有老模式记录）\n3. 携程侧该用户在授信环节有已签署的协议文件',
  '1. 调用携程授信进件提交接口\n2. 查询aps_stg3.partner_agreement_instance表\n3. 查询lps日志确认是否调用ags协议服务',
  '1. 授信进件提交成功\n2. partner_agreement_instance表落库记录，status=SUCCESS, file_id不为空\n3. lps日志中无调用ags协议服务的记录',
  '高','是','携程,新模式,授信,新户,协议拉取')

c('JYSG-149007',
  '携程新模式-授信进件提交-协议缺失触发拒绝',
  '验证enabled=Y下，授信进件提交时携程侧无用户已签署协议，触发交易拒绝，不写user_order_notice_info',
  '1. aps_stg3.partner_agreement_config：enabled=Y, grayscale_ratio>1000\n2. 携程渠道新用户，但携程侧该用户在授信环节无已签署的协议文件',
  '1. 调用携程授信进件提交接口\n2. 查询aps.order_info表\n3. 查询user_order_notice_info表',
  '1. 授信进件提交成功但交易被拒绝\n2. aps.order_info状态变更为ARJ（拒绝）\n3. user_order_notice_info表无记录',
  '高','否','携程,新模式,授信,协议缺失,拒绝')

c('JYSG-149007',
  '携程新模式-授信进件提交-协议拉取超时异常',
  '验证enabled=Y下，协议拉取超时(超过timeout_minutes)，触发拒绝',
  '1. aps_stg3.partner_agreement_config：enabled=Y, timeout_minutes设为较短值\n2. 携程渠道新用户，模拟协议拉取超时',
  '1. 修改partner_agreement_config.timeout_minutes为2\n2. 调用携程授信进件提交接口（协议拉取全部超时）\n3. 查询aps.order_info表',
  '1. 授信进件提交成功但交易被拒绝\n2. aps.order_info状态变更为ARJ',
  '中','否','携程,新模式,授信,超时,异常')

c('JYSG-149007',
  '携程新模式-授信进件-协议补拉重试成功',
  '验证协议首次拉取失败后，定时任务补拉重试成功（fast_retry_window内快速重试）',
  '1. partner_agreement_config：enabled=Y, fast_retry_window_minutes=10, fast_retry_interval_minutes=1\n2. 存在partner_agreement_instance表中status=FAIL的记录',
  '1. 确认partner_agreement_instance表有retry_count<上限的FAIL记录\n2. 等待定时任务触发补拉\n3. 查询partner_agreement_instance表',
  '1. 定时任务触发补拉\n2. 补拉成功后partner_agreement_instance.status变为SUCCESS, file_id不为空',
  '中','否','携程,新模式,授信,补拉,重试')

c('JYSG-149007',
  '携程新模式-授信通过-授信通知触发协议下载',
  '验证授信状态通过后，aps收到授信通知并调用downLoadPartnerAgreement接口下载协议文件',
  '1. partner_agreement_config：enabled=Y\n2. 授信进件已提交成功且协议拉取成功',
  '1. 触发授信状态通过通知\n2. 查询user_order_notice_info表\n3. 查询aps-app日志确认downLoadPartnerAgreement调用',
  '1. user_order_notice_info表生成一条记录\n2. aps-app日志中有调用downLoadPartnerAgreement接口的记录',
  '中','否','携程,新模式,授信,授信通知,协议下载')

c('JYSG-149007',
  '携程新模式-授信拒绝-不触发协议下载',
  '验证授信状态被拒绝时，不触发协议下载通知',
  '1. partner_agreement_config：enabled=Y\n2. 授信进件已提交，模拟授信拒绝',
  '1. 触发授信状态拒绝\n2. 查询user_order_notice_info表\n3. 查询aps-app日志',
  '1. user_order_notice_info表生成一条记录（记录拒绝结果）\n2. aps-app日志中无downLoadPartnerAgreement调用',
  '中','否','携程,新模式,授信,授信拒绝')

c('JYSG-149007',
  '携程新模式-授信通过-协议下载失败可补拉',
  '验证授信通过但协议下载失败时，记录正常生成，后续补拉可成功',
  '1. partner_agreement_config：enabled=Y\n2. 授信进件已提交，模拟协议下载失败',
  '1. 触发授信状态通过\n2. 模拟downLoadPartnerAgreement调用失败\n3. 查询partner_agreement_instance表retry_count\n4. 触发补拉定时任务',
  '1. partner_agreement_instance.status=FAIL, retry_count递增\n2. 补拉定时任务重试后status变为SUCCESS',
  '低','否','携程,新模式,授信,协议下载失败,补拉')

c('JYSG-149007',
  '携程新模式-授信进件-协议拉取成功授信拒绝',
  '验证协议拉取成功但后续授信被拒绝时，已保存的协议记录不受影响',
  '1. partner_agreement_config：enabled=Y\n2. 授信进件协议拉取成功，模拟后续授信拒绝',
  '1. 调用授信进件提交接口（协议拉取成功）\n2. 模拟授信拒绝\n3. 查询partner_agreement_instance表',
  '1. partner_agreement_instance表协议记录保持SUCCESS\n2. 授信拒绝不影响已保存的协议记录',
  '低','否','携程,新模式,授信,协议成功,授信拒绝')

# ============ A2. 新模式 - 授信进件 老户(360API重授信) (2条) ============
c('JYSG-149007',
  '携程新模式-授信进件提交-360API老户重授信-协议拉取成功',
  '验证360API老链路用户在partner_agreement_mode_record中有老模式记录时，授信进件拉取协议成功，检查APS日志无ags协议调用',
  '1. partner_agreement_config：enabled=Y\n2. 携程渠道老用户（partner_agreement_mode_record表中有该用户的老模式记录）\n3. 携程侧该用户在授信环节有已签署的协议文件',
  '1. 通过360API老链路调用授信进件提交接口\n2. 查询partner_agreement_instance表\n3. 查看APS日志确认是否调用ags协议服务',
  '1. 授信进件提交成功\n2. partner_agreement_instance表落库记录，status=SUCCESS, file_id不为空\n3. APS日志中无调用ags协议服务的记录',
  '中','否','携程,新模式,授信,360API,老户,协议拉取')

c('JYSG-149007',
  '携程新模式-授信进件提交-360API老户-协议缺失触发拒绝',
  '验证360API老户重授信时协议缺失，触发交易拒绝',
  '1. partner_agreement_config：enabled=Y\n2. 携程渠道老用户（partner_agreement_mode_record中有老模式记录），但携程侧无已签署协议',
  '1. 通过360API老链路调用授信进件提交接口\n2. 查询aps.order_info表\n3. 查询user_order_notice_info表',
  '1. 授信进件提交成功但交易被拒绝\n2. aps.order_info状态变更为ARJ（拒绝）\n3. user_order_notice_info表无记录',
  '中','否','携程,新模式,授信,360API,老户,协议缺失,拒绝')

# ============ B. 新模式 - 借款提交 (8条) ============
c('JYSG-149008',
  '携程新模式-借款提交-协议拉取成功',
  '验证enabled=Y下，借款提交时成功拉取携程侧用户章协议（含协议编号传递），检查tfs日志无ags协议调用',
  '1. partner_agreement_config：enabled=Y\n2. 授信已完成，用户有借款额度\n3. 携程侧已生成借款协议且有用户章',
  '1. 调用携程借款提交接口（含协议编号参数）\n2. 查询aps_stg3.partner_agreement_instance表\n3. 查询aps.order_iou表\n4. 查看tfs日志确认是否调用ags协议服务',
  '1. 借款提交成功\n2. partner_agreement_instance表落库借款协议记录，status=SUCCESS, contract_no不为空\n3. aps.order_iou表有借款记录\n4. tfs日志中无调用ags协议服务的记录',
  '高','是','携程,新模式,借款,协议拉取,协议编号')

c('JYSG-149008',
  '携程新模式-借款提交-协议缺失触发拒绝',
  '验证enabled=Y下，借款提交时携程侧无已签署协议，触发拒绝',
  '1. partner_agreement_config：enabled=Y\n2. 授信已完成，但携程侧无借款环节协议文件',
  '1. 调用携程借款提交接口\n2. 查询aps.order_iou表\n3. 查询user_order_notice_info表',
  '1. 借款提交成功但交易被拒绝\n2. aps.order_iou表state为9（拒绝）\n3. user_order_notice_info表无记录',
  '高','否','携程,新模式,借款,协议缺失,拒绝')

c('JYSG-149008',
  '携程新模式-借款提交-协议拉取超时异常',
  '验证enabled=Y下，借款提交协议拉取超时(超过timeout_minutes)，触发拒绝',
  '1. partner_agreement_config：enabled=Y, timeout_minutes设为较短值\n2. 授信已完成，模拟协议拉取超时',
  '1. 修改partner_agreement_config.timeout_minutes为2\n2. 调用携程借款提交接口（协议拉取全部超时）\n3. 查询aps.order_iou表',
  '1. 借款提交成功但交易被拒绝\n2. aps.order_iou表state为9（拒绝）',
  '中','否','携程,新模式,借款,超时,异常')

c('JYSG-149008',
  '携程新模式-借款通过-放款通知触发协议下载',
  '验证借款放款通过后，aps收到通知并调用downLoadPartnerAgreement接口下载协议文件',
  '1. partner_agreement_config：enabled=Y\n2. 借款已提交成功且协议拉取成功',
  '1. 触发借款放款通过通知\n2. 查询user_order_notice_info表\n3. 查询aps-app日志确认downLoadPartnerAgreement接口调用',
  '1. user_order_notice_info表生成一条记录\n2. aps-app日志中有调用downLoadPartnerAgreement接口的记录',
  '中','否','携程,新模式,借款,放款通知,协议下载')

c('JYSG-149008',
  '携程新模式-借款放款拒绝-不触发协议下载',
  '验证借款放款被拒绝时，不触发协议下载通知',
  '1. partner_agreement_config：enabled=Y\n2. 借款已提交，模拟放款拒绝',
  '1. 触发借款放款拒绝\n2. 查询user_order_notice_info表\n3. 查询aps-app日志',
  '1. user_order_notice_info表生成一条记录（记录拒绝结果）\n2. aps-app日志中无downLoadPartnerAgreement调用',
  '中','否','携程,新模式,借款,放款拒绝')

c('JYSG-149008',
  '携程新模式-借款通过-协议下载失败可补拉',
  '验证借款放款通过但协议下载失败时，记录正常生成，支持后续补拉',
  '1. partner_agreement_config：enabled=Y\n2. 借款已提交，模拟协议下载失败',
  '1. 触发借款放款通过\n2. 模拟downLoadPartnerAgreement调用失败\n3. 查询partner_agreement_instance表retry_count\n4. 触发协议补拉定时任务',
  '1. partner_agreement_instance.status=FAIL, retry_count递增\n2. 补拉定时任务重试后status变为SUCCESS',
  '低','否','携程,新模式,借款,协议下载失败,补拉')

c('JYSG-149008',
  '携程新模式-借款提交-协议拉取成功放款拒绝',
  '验证借款协议拉取成功但后续放款拒绝时，已保存的协议记录不受影响',
  '1. partner_agreement_config：enabled=Y\n2. 借款协议拉取成功，模拟后续放款拒绝',
  '1. 调用借款提交接口（协议拉取成功）\n2. 模拟放款拒绝\n3. 查询partner_agreement_instance表',
  '1. partner_agreement_instance表协议记录保持SUCCESS\n2. 放款拒绝不影响已保存的协议记录',
  '低','否','携程,新模式,借款,协议成功,放款拒绝')

c('JYSG-149008',
  '携程新模式-借款提交-协议编号传递验证',
  '验证借款提交时携程侧协议编号(contract_no)随借款申请接口推送至APS并正确保存',
  '1. partner_agreement_config：enabled=Y\n2. 携程侧已生成借款协议且含协议编号',
  '1. 调用携程借款提交接口（含协议编号参数）\n2. 查询partner_agreement_instance表contract_no字段\n3. 查询合同系统确认协议编号已上送',
  '1. 借款提交成功\n2. partner_agreement_instance.contract_no不为空\n3. 合同系统收到协议编号并保存',
  '高','是','携程,新模式,借款,协议编号')

# ============ C. 新模式 - 特殊场景 (2条) ============
c('JYSG-149008',
  '携程新模式-借款发送转出中-协议缺失拒绝',
  '验证借款提交后状态为转出中时，协议缺失触发拒绝',
  '1. partner_agreement_config：enabled=Y\n2. 借款提交后订单状态为转出中，协议未拉取完成',
  '1. 提交借款，触发转出中状态\n2. 查询aps.order_iou表',
  '1. 借款提交成功但被拒绝\n2. aps.order_iou表state为9（拒绝）',
  '中','否','携程,新模式,借款,转出中,拒绝')

c('JYSG-149008,JYSG-149009',
  '携程新模式-借款节点聚合流程-放款与协议联动',
  '验证借款环节聚合流程中放款与协议下载的正确联动',
  '1. partner_agreement_config：enabled=Y\n2. 借款提交完成，需走过放款+协议下载聚合流程',
  '1. 执行借款聚合流程（放款+协议下载）\n2. 查询user_order_notice_info表\n3. 查询partner_agreement_instance表',
  '1. 借款聚合流程执行完成\n2. 放款+协议下载联动正确\n3. partner_agreement_instance表协议记录status=SUCCESS',
  '中','否','携程,新模式,借款,聚合流程')

# ============ D. 新模式 - 签署完成拉取协议 (4条) ============
c('JYSG-149009',
  '携程新模式-用户签署完成后拉取已签章协议',
  '验证用户完成签署后，系统通过partner_agreement_instance拉取含用户章的协议文件成功',
  '1. partner_agreement_config：enabled=Y\n2. 放款通过，用户已签署协议',
  '1. 触发用户签署完成\n2. 调用协议拉取接口\n3. 查询partner_agreement_instance表file_id字段',
  '1. 协议拉取成功\n2. partner_agreement_instance.file_id不为空\n3. 协议文件含用户电子签章，内容可预览',
  '高','是','携程,新模式,签章,协议拉取')

c('JYSG-149009',
  '携程新模式-签署完成后-放款拒绝仍可拉取协议',
  '验证放款被拒绝但用户已签署协议时，仍可拉取协议文件',
  '1. partner_agreement_config：enabled=Y\n2. 放款被拒绝，但用户已签署协议',
  '1. 触发放款拒绝\n2. 调用协议拉取接口\n3. 查询partner_agreement_instance表',
  '1. 协议拉取成功\n2. partner_agreement_instance.file_id不为空',
  '中','否','携程,新模式,签章,放款拒绝,协议拉取')

c('JYSG-149009',
  '携程新模式-签署完成后-协议拉取失败可重试',
  '验证放款通过但协议拉取失败时，用户可重新拉取',
  '1. partner_agreement_config：enabled=Y\n2. 放款通过，模拟协议拉取失败',
  '1. 触发放款通过\n2. 首次协议拉取失败\n3. 重新调用协议拉取接口',
  '1. 首次拉取失败，partner_agreement_instance.status=FAIL\n2. 重新拉取成功，status变为SUCCESS',
  '低','否','携程,新模式,签章,协议拉取失败,重试')

c('JYSG-149009',
  '携程新模式-签署完成后-协议拉取成功放款失败',
  '验证协议拉取成功但后续放款失败时，已拉取的协议文件不受影响',
  '1. partner_agreement_config：enabled=Y\n2. 协议拉取成功，模拟后续放款失败',
  '1. 协议拉取成功\n2. 模拟放款失败\n3. 验证协议文件完整性',
  '1. 已拉取的协议文件不受放款失败影响\n2. partner_agreement_instance.file_id保持完整',
  '低','否','携程,新模式,签章,协议成功,放款失败')

# ============ E. 高登模式 (2条) ============
c('JYSG-149007',
  '携程高登模式-进件提交-不触发协议拉取',
  '验证partner_agreement_config.enabled=N + grayscale_ratio>1000下，授信进件不触发新模式协议拉取，走原有ags协议流程',
  '1. aps_stg3.partner_agreement_config：enabled=N, grayscale_ratio>1000（灰度关闭）\n2. 携程渠道测试账号',
  '1. 调用携程授信进件提交接口\n2. 查询user_order_notice_info表\n3. 查看lps日志确认是否调用ags协议服务\n4. 查询partner_agreement_instance表',
  '1. 授信进件提交成功，走原有流程\n2. user_order_notice_info表无记录\n3. lps日志中有调用ags协议服务的记录\n4. partner_agreement_instance表无新增记录',
  '高','是','携程,高登模式,进件,不触发协议')

c('JYSG-149008',
  '携程高登模式-借款提交-不触发协议拉取',
  '验证enabled=N + grayscale_ratio>1000下，借款提交不触发新模式协议拉取，走原有ags协议流程',
  '1. partner_agreement_config：enabled=N, grayscale_ratio>1000\n2. 授信进件已完成',
  '1. 调用携程借款提交接口\n2. 查询user_order_notice_info表\n3. 查看tfs日志确认是否调用ags协议服务\n4. 查询partner_agreement_instance表',
  '1. 借款提交成功，走原有流程\n2. user_order_notice_info表无记录\n3. tfs日志中有调用ags协议服务的记录\n4. partner_agreement_instance表无新增记录',
  '高','是','携程,高登模式,借款,不触发协议')

# ============ F. 灰度/白名单控制 (5条) ============
c('JYSG-149011',
  '携程灰度模式-enabled=Y+grayscale_ratio控制灰度比例',
  '验证enabled=Y, grayscale_ratio在0~1000区间按比例灰度，命中灰度的用户走新模式',
  '1. partner_agreement_config：enabled=Y, grayscale_ratio=500（50%灰度）\n2. 命中灰度比例的携程用户账号',
  '1. 以灰度命中用户调用授信进件提交\n2. 查询partner_agreement_instance表\n3. 查看lps日志确认ags协议调用情况',
  '1. 灰度命中用户走新模式\n2. partner_agreement_instance表有协议记录\n3. lps日志中无ags协议调用',
  '中','否','携程,灰度,grayscale_ratio,新模式')

c('JYSG-149011',
  '携程灰度模式-未命中灰度用户走高登模式',
  '验证enabled=Y, grayscale_ratio=500但用户未命中灰度时，走高登模式',
  '1. partner_agreement_config：enabled=Y, grayscale_ratio=500\n2. 未命中灰度比例的携程用户账号',
  '1. 以未命中灰度用户调用授信进件提交\n2. 查询partner_agreement_instance表\n3. 查看lps日志',
  '1. 未命中灰度用户走高登模式\n2. partner_agreement_instance表无记录\n3. lps日志中有ags协议调用',
  '中','否','携程,灰度,grayscale_ratio,高登模式')

c('JYSG-149011',
  '携程灰度模式-手机号白名单用户走新模式',
  '验证mobile_whitelist中配置的手机号对应用户走新模式',
  '1. partner_agreement_config：enabled=Y, grayscale_ratio>1000, mobile_whitelist含指定手机号\n2. 该手机号对应的携程用户账号',
  '1. 以白名单手机号用户调用授信进件提交\n2. 以同用户调用借款提交\n3. 查询partner_agreement_instance表',
  '1. 授信进件和借款提交均走新模式\n2. partner_agreement_instance表有协议记录\n3. 流程中不触发ags协议服务',
  '中','否','携程,灰度,白名单,手机号')

c('JYSG-149011',
  '携程灰度模式-身份证号白名单用户走新模式',
  '验证id_number_whitelist中配置的身份证号对应用户走新模式',
  '1. partner_agreement_config：enabled=Y, grayscale_ratio>1000, id_number_whitelist含指定身份证号\n2. 该身份证号对应的用户',
  '1. 以白名单身份证号用户调用借款提交\n2. 查询partner_agreement_instance表',
  '1. 借款提交走新模式\n2. partner_agreement_instance表有协议记录',
  '低','否','携程,灰度,白名单,身份证号')

c('JYSG-149011',
  '携程灰度模式-order_no白名单订单走新模式',
  '验证order_no_whitelist中配置的订单走新模式',
  '1. partner_agreement_config：enabled=Y, grayscale_ratio>1000, order_no_whitelist含指定order_no\n2. 该order_no对应的订单',
  '1. 以白名单order_no调用借款提交\n2. 查询partner_agreement_instance表',
  '1. 该order_no走新模式\n2. partner_agreement_instance表有协议记录',
  '低','否','携程,灰度,白名单,order_no')

# ===== Write Excel =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

for ci, h in enumerate(hdrs, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = cwrap; cell.border = border

for ri, case in enumerate(cases, 2):
    for ci, val in enumerate(case, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.font = cfont; cell.alignment = wrap; cell.border = border
        if ci in (1, 9, 10, 11, 12, 13): cell.alignment = cwrap

widths = [24, 28, 40, 48, 56, 60, 60, 18, 8, 8, 10, 12, 10, 32]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.row_dimensions[1].height = 28
for r in range(2, len(cases)+2): ws.row_dimensions[r].height = 80
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:N{len(cases)+1}'

out = r'D:\测试用例\20260611测试用例\携程协议_优化版.xlsx'
wb.save(out)
print(f'Done: {out}')
print(f'Total: {len(cases)} cases')
print(f'  高={sum(1 for c in cases if c[8]=="高")}, 中={sum(1 for c in cases if c[8]=="中")}, 低={sum(1 for c in cases if c[8]=="低")}')
print(f'  回归=是: {sum(1 for c in cases if c[11]=="是")}')
print(f'  覆盖: 授信新户={sum(1 for c in cases if "新户" in c[2])}, 授信老户={sum(1 for c in cases if "老户" in c[2])}, 借款={sum(1 for c in cases if "借款" in c[2])}, 签署={sum(1 for c in cases if "签署" in c[2])}, 高登={sum(1 for c in cases if "高登" in c[2])}, 灰度={sum(1 for c in cases if "灰度" in c[2])}')
