import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Styles
header_font_white = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
cell_font = Font(name='微软雅黑', size=10)
wrap = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ['STORY编号', '节点路径', '用例名称', '摘要', '前置条件', '执行步骤', '预期结果', '实际结果', '优先级', '作者', '类型', '是否需要回归', '应用', '标签']

# ===== OPTIMIZED TEST CASES (30 cases) =====
cases = []

def add(story, node, name, summary, precond, steps, expected, priority, regression, tags, author='王雪', app='aps-app', case_type='功能测试'):
    cases.append({
        'story': story, 'node': node, 'name': name, 'summary': summary,
        'precondition': precond, 'steps': steps, 'expected': expected,
        'priority': priority, 'author': author, 'type': case_type,
        'regression': regression, 'app': app, 'tags': tags
    })

NODE = '/全部需求/产品需求/API/携程/协议'

# ===== A. 新模式 - 授信进件 (8条) =====
add('JYSG-149007', NODE,
    '携程新模式-授信进件提交-协议拉取成功',
    '验证新模式(partner_agreement_config开关打开)下，授信进件提交时成功拉取携程侧用户章协议文件',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 携程渠道测试账号，该用户在授信环节有已签署的协议文件',
    '1. 调用携程授信进件提交接口\n2. 查询partner_agreement_instance表\n3. 查看APS日志确认协议拉取请求',
    '1. 授信进件提交成功，接口返回成功\n2. partner_agreement_instance表落库记录，fild_id和status字段均为success\n3. APS日志中无调用ags协议服务的记录（协议由携程侧生成）',
    '高', '是', '携程,新模式,授信,协议拉取')

add('JYSG-149007', NODE,
    '携程新模式-授信进件提交-协议缺失触发拒绝',
    '验证新模式(开关打开)下，授信进件提交时若携程侧无用户已签署的协议文件，触发交易拒绝',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 携程渠道测试账号，但该用户在授信环节无已签署的协议文件',
    '1. 调用携程授信进件提交接口\n2. 查询aps.order_info表\n3. 查询user_order_notice_info表',
    '1. 授信进件提交成功但交易被拒绝\n2. aps.order_info状态变更为ARJ（拒绝）\n3. user_order_notice_info表无记录（不发送通知）',
    '高', '否', '携程,新模式,授信,协议缺失,拒绝')

add('JYSG-149007', NODE,
    '携程新模式-授信进件提交-协议拉取超时异常',
    '验证新模式(开关打开)下，授信进件提交时若协议拉取超时（超过timeout_minutes），触发拒绝',
    '1. Apollo配置：partner_agreement_config开关打开，timeout_minutes配置为指定时长\n2. 携程渠道测试账号，模拟协议拉取超时场景',
    '1. 修改Apollo配置：partner_agreement_config.timeout_minutes设为较短时长\n2. 调用携程授信进件提交接口（协议拉取全部超时）\n3. 查询aps.order_info表\n4. 查询user_order_notice_info表',
    '1. 授信进件提交成功但交易被拒绝\n2. aps.order_info状态变更为ARJ（拒绝）\n3. user_order_notice_info表无记录',
    '中', '否', '携程,新模式,授信,超时,异常')

add('JYSG-149007', NODE,
    '携程新模式-授信进件-协议补拉重试成功',
    '验证协议拉取失败后，定时任务补拉重试成功',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 存在之前协议拉取失败但后续携程侧已生成协议的测试数据',
    '1. 确认数据库中存在协议拉取失败的记录\n2. 触发协议补拉定时任务\n3. 查询partner_agreement_instance表',
    '1. 定时任务执行成功\n2. 补拉成功后partner_agreement_instance表记录更新，fild_id和status为success',
    '中', '否', '携程,新模式,授信,补拉,重试')

add('JYSG-149007', NODE,
    '携程新模式-授信进件-放款通知imgp拉取协议',
    '验证授信进件放款通过后，imgp收到通知并拉取协议文件',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 授信进件已提交成功且协议拉取成功',
    '1. 触发放款通知\n2. 查询user_order_notice_info表\n3. 检查imgp日志确认协议拉取',
    '1. user_order_notice_info表生成一条记录\n2. 定时任务触发imgp拉取协议成功\n3. partner_agreement_instance表状态更新',
    '中', '否', '携程,新模式,授信,放款通知,imgp')

add('JYSG-149007', NODE,
    '携程新模式-授信进件-放款拒绝场景',
    '验证授信进件放款被拒绝时，不触发imgp协议拉取通知',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 授信进件已提交，模拟放款拒绝',
    '1. 触发放款拒绝\n2. 查询user_order_notice_info表\n3. 检查imgp日志',
    '1. user_order_notice_info表生成一条记录（记录拒绝结果）\n2. 定时任务通知imgp时不触发协议拉取',
    '中', '否', '携程,新模式,授信,放款拒绝')

add('JYSG-149007', NODE,
    '携程新模式-授信进件-放款成功协议拉取失败',
    '验证放款成功但imgp拉取协议失败时，通知记录正常生成，支持后续补拉',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 授信进件已提交，模拟imgp协议拉取失败',
    '1. 触发放款通过\n2. 查询user_order_notice_info表\n3. 检查imgp日志确认协议拉取失败\n4. 触发协议补拉定时任务',
    '1. user_order_notice_info表记录正常生成（放款成功）\n2. imgp协议拉取失败，日志记录异常\n3. 协议补拉定时任务可重试并成功',
    '低', '否', '携程,新模式,授信,放款成功,协议拉取失败')

add('JYSG-149007', NODE,
    '携程新模式-授信进件-协议拉取成功放款失败',
    '验证协议拉取成功但放款失败时，协议记录保持完整不受影响',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 授信进件协议拉取成功，模拟后续放款失败',
    '1. 调用授信进件提交接口（协议拉取成功）\n2. 模拟放款失败\n3. 查询partner_agreement_instance表',
    '1. partner_agreement_instance表协议记录保持success\n2. 放款失败不影响已保存的协议记录',
    '低', '否', '携程,新模式,授信,协议成功,放款失败')

# ===== B. 新模式 - 借款提交 (8条) =====
add('JYSG-149008', NODE,
    '携程新模式-借款提交-协议拉取成功',
    '验证新模式(开关打开)下，借款提交时成功拉取携程侧用户章协议文件（含协议编号传递）',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 授信进件已完成，用户有借款额度\n3. 携程侧已生成借款协议且有用户章',
    '1. 调用携程借款提交接口（含协议编号参数）\n2. 查询partner_agreement_instance表\n3. 查询aps.order_iou表\n4. 查看APS日志确认协议拉取请求',
    '1. 借款提交成功，接口返回成功\n2. partner_agreement_instance表落库借款协议记录，fild_id和status为success\n3. aps.order_iou表有借款记录\n4. APS日志中无调用ags协议服务的记录',
    '高', '是', '携程,新模式,借款,协议拉取,协议编号')

add('JYSG-149008', NODE,
    '携程新模式-借款提交-协议缺失触发拒绝',
    '验证新模式(开关打开)下，借款提交时若携程侧无用户已签署的协议文件，触发交易拒绝',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 授信进件已完成，但携程侧无借款环节协议文件',
    '1. 调用携程借款提交接口\n2. 查询aps.order_iou表\n3. 查询user_order_notice_info表',
    '1. 借款提交成功但交易被拒绝\n2. aps.order_iou表状态为9（拒绝）\n3. user_order_notice_info表无记录',
    '高', '否', '携程,新模式,借款,协议缺失,拒绝')

add('JYSG-149008', NODE,
    '携程新模式-借款提交-协议拉取超时异常',
    '验证新模式(开关打开)下，借款提交时若协议拉取超时（超过timeout_minutes），触发拒绝',
    '1. Apollo配置：partner_agreement_config开关打开，timeout_minutes设为较短时长\n2. 授信进件已完成，模拟协议拉取超时',
    '1. 修改Apollo配置：partner_agreement_config.timeout_minutes设为较短时长\n2. 调用携程借款提交接口（协议拉取全部超时）\n3. 查询aps.order_iou表\n4. 查询user_order_notice_info表',
    '1. 借款提交成功但交易被拒绝\n2. aps.order_iou表状态为9（拒绝）\n3. user_order_notice_info表无记录',
    '中', '否', '携程,新模式,借款,超时,异常')

add('JYSG-149008', NODE,
    '携程新模式-借款提交-放款通知imgp拉取协议',
    '验证借款放款通过后，imgp收到通知并拉取协议文件',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 借款已提交成功且协议拉取成功',
    '1. 触发借款放款通知\n2. 查询user_order_notice_info表\n3. 检查imgp日志确认协议拉取',
    '1. user_order_notice_info表生成一条记录\n2. 定时任务触发imgp拉取协议成功',
    '中', '否', '携程,新模式,借款,放款通知,imgp')

add('JYSG-149008', NODE,
    '携程新模式-借款提交-放款拒绝场景',
    '验证借款放款被拒绝时，不触发imgp协议拉取通知',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 借款已提交，模拟放款拒绝',
    '1. 触发借款放款拒绝\n2. 查询user_order_notice_info表\n3. 检查imgp日志',
    '1. user_order_notice_info表生成一条记录（记录拒绝结果）\n2. 定时任务通知imgp时不触发协议拉取',
    '中', '否', '携程,新模式,借款,放款拒绝')

add('JYSG-149008', NODE,
    '携程新模式-借款提交-放款成功协议拉取失败',
    '验证借款放款成功但imgp拉取协议失败时，记录正常生成，支持后续补拉',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 借款已提交，模拟imgp协议拉取失败',
    '1. 触发借款放款通过\n2. 查询user_order_notice_info表\n3. 检查imgp日志确认协议拉取失败\n4. 触发协议补拉定时任务',
    '1. user_order_notice_info表记录正常生成（放款成功）\n2. imgp协议拉取失败，日志记录异常\n3. 协议补拉定时任务可重试并成功',
    '低', '否', '携程,新模式,借款,放款成功,协议拉取失败')

add('JYSG-149008', NODE,
    '携程新模式-借款提交-协议拉取成功放款失败',
    '验证借款协议拉取成功但放款失败时，协议记录保持完整不受影响',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 借款协议拉取成功，模拟后续放款失败',
    '1. 调用借款提交接口（协议拉取成功）\n2. 模拟放款失败\n3. 查询partner_agreement_instance表',
    '1. partner_agreement_instance表协议记录保持success\n2. 放款失败不影响已保存的协议记录',
    '低', '否', '携程,新模式,借款,协议成功,放款失败')

add('JYSG-149008', NODE,
    '携程新模式-借款提交-协议编号传递验证',
    '验证借款提交时携程侧协议编号随借款申请接口同步推送至APS并正确保存',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 携程侧已生成借款协议且含协议编号',
    '1. 调用携程借款提交接口（含协议编号参数）\n2. 查询partner_agreement_instance表验证协议编号字段\n3. 查询合同系统确认协议编号已上送',
    '1. 借款提交成功\n2. partner_agreement_instance表协议编号字段不为空\n3. 合同系统收到协议编号并保存',
    '高', '是', '携程,新模式,借款,协议编号')

# ===== C. 新模式 - 特殊场景 (3条) =====
add('JYSG-149008', NODE,
    '携程新模式-借款发送转出中-拒绝',
    '验证借款提交后状态为转出中时，协议缺失触发拒绝',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 借款提交后订单状态为转出中，协议未拉取完成',
    '1. 提交借款，触发转出中状态\n2. 查询aps.order_iou表',
    '1. 借款提交成功但被拒绝\n2. aps.order_iou表state为9（拒绝）',
    '中', '否', '携程,新模式,借款,转出中,拒绝')

add('JYSG-149008,JYSG-149009', NODE,
    '携程新模式-借款节点聚合流程',
    '验证借款环节聚合流程中协议处理与放款的正确联动',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 借款提交完成，需走过放款+协议拉取聚合流程',
    '1. 执行借款聚合流程（放款+协议拉取）\n2. 查询user_order_notice_info表\n3. 查询partner_agreement_instance表',
    '1. 借款聚合流程执行完成\n2. 放款+协议拉取联动正确\n3. partner_agreement_instance表协议记录状态为success',
    '中', '否', '携程,新模式,借款,聚合流程')

add('JYSG-149010', NODE,
    '携程新模式-借款流程-明示表开关控制',
    '验证明示表(综合融资成本明示表)开关关闭时不阻塞借款流程，开关打开后触发协议拉取',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 明示表开关先关闭后打开',
    '1. 明示表开关关闭：调用借款提交接口\n2. 查询借款是否正常提交\n3. 明示表开关打开：再次调用借款提交接口\n4. 查询是否触发明示表协议拉取',
    '1. 开关关闭时借款正常提交，不拉取明示表\n2. 开关打开后借款提交触发明示表协议拉取\n3. 不影响其他协议的正常拉取',
    '中', '否', '携程,新模式,借款,明示表,开关')

# ===== D. 新模式 - 签署完成拉取协议 (4条) =====
add('JYSG-149009', NODE,
    '携程新模式-用户签署完成后拉取已签章协议',
    '验证用户完成签署后，系统拉取带有用户章的协议文件成功',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 放款通过，用户已签署协议',
    '1. 触发放款通过后用户签署完成\n2. 调用协议拉取接口\n3. 验证拉取的协议文件含用户章',
    '1. 协议拉取成功，返回协议文件\n2. 协议文件含用户电子签章，内容可预览',
    '高', '是', '携程,新模式,签章,协议拉取')

add('JYSG-149009', NODE,
    '携程新模式-签署完成后-放款拒绝仍可拉取协议',
    '验证放款被拒绝但用户已签署协议时，仍可拉取协议文件',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 放款被拒绝，但用户已签署协议',
    '1. 触发放款拒绝\n2. 调用协议拉取接口',
    '1. 协议拉取成功，返回协议文件\n2. 协议文件内容可预览',
    '中', '否', '携程,新模式,签章,放款拒绝,协议拉取')

add('JYSG-149009', NODE,
    '携程新模式-签署完成后-放款成功协议拉取失败可重试',
    '验证放款成功但协议拉取失败时，用户可重新拉取',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 放款通过，模拟协议拉取失败',
    '1. 触发放款通过\n2. 首次协议拉取失败\n3. 重新调用协议拉取接口',
    '1. 首次拉取失败，记录失败原因\n2. 重新拉取成功，返回协议文件',
    '低', '否', '携程,新模式,签章,协议拉取失败,重试')

add('JYSG-149009', NODE,
    '携程新模式-签署完成后-协议拉取成功放款失败',
    '验证协议拉取成功但放款失败时，已拉取的协议文件不受影响',
    '1. Apollo配置：partner_agreement_config开关打开\n2. 协议拉取成功，模拟后续放款失败',
    '1. 协议拉取成功\n2. 模拟放款失败\n3. 验证协议文件完整性',
    '1. 已拉取的协议文件不受放款失败影响\n2. 协议文件内容和签章完整',
    '低', '否', '携程,新模式,签章,协议成功,放款失败')

# ===== E. 高登模式 (2条) =====
add('JYSG-149007', NODE,
    '携程高登模式-进件提交-不触发协议拉取',
    '验证高登模式(partner_agreement_config开关关闭且未开启灰度/白名单)下，授信进件提交不触发新模式协议拉取，走原有ags协议流程',
    '1. Apollo配置：partner_agreement_config开关关闭\n2. 灰度开关关闭，白名单未配置\n3. 携程渠道测试账号',
    '1. 调用携程授信进件提交接口\n2. 查询user_order_notice_info表\n3. 查看lps日志确认是否调用ags协议服务',
    '1. 授信进件提交成功，走原有流程\n2. user_order_notice_info表无记录（不发送通知）\n3. lps日志中有调用ags协议服务的记录（原有流程）\n4. partner_agreement_instance表无新增记录',
    '高', '是', '携程,高登模式,进件,不触发协议')

add('JYSG-149008', NODE,
    '携程高登模式-借款提交-不触发协议拉取',
    '验证高登模式(开关关闭且未开启灰度)下，借款提交不触发新模式协议拉取，走原有ags协议流程',
    '1. Apollo配置：partner_agreement_config开关关闭\n2. 灰度开关关闭，白名单未配置\n3. 授信进件已完成',
    '1. 调用携程借款提交接口\n2. 查询user_order_notice_info表\n3. 查看tfs日志确认是否调用ags协议服务',
    '1. 借款提交成功，走原有流程\n2. user_order_notice_info表无记录\n3. tfs日志中有调用ags协议服务的记录（原有流程）\n4. partner_agreement_instance表无新增记录',
    '高', '是', '携程,高登模式,借款,不触发协议')

# ===== F. 灰度/白名单控制 (5条) =====
add('JYSG-149011', NODE,
    '携程灰度模式-白名单用户走新模式',
    '验证灰度模式下，白名单用户走新模式触发协议拉取',
    '1. Apollo配置：partner_agreement_config开关关闭\n2. 灰度白名单开关打开，白名单配置为指定用户\n3. 白名单用户账号（如手机号在配置列表中）',
    '1. 以白名单用户身份调用授信进件提交接口\n2. 查询user_order_notice_info表\n3. 检查是否触发新模式协议拉取',
    '1. 白名单用户走新模式\n2. user_order_notice_info表有记录（新模式通知）\n3. partner_agreement_instance表有协议记录',
    '中', '否', '携程,灰度,白名单,新模式')

add('JYSG-149011', NODE,
    '携程灰度模式-非白名单用户走高登模式',
    '验证灰度模式下，非白名单用户走高登模式不触发协议拉取',
    '1. Apollo配置：partner_agreement_config开关关闭\n2. 灰度白名单开关打开，白名单不含该用户\n3. 非白名单用户账号',
    '1. 以非白名单用户身份调用授信进件提交接口\n2. 查询user_order_notice_info表\n3. 检查是否走原有高登流程',
    '1. 用户走高登模式\n2. user_order_notice_info表无记录\n3. 走原有ags协议流程',
    '中', '否', '携程,灰度,白名单,高登模式')

add('JYSG-149011', NODE,
    '携程灰度模式-手机号维度白名单控制',
    '验证以手机号为白名单维度的灰度控制，用户走新模式',
    '1. Apollo配置：partner_agreement_config开关关闭\n2. 灰度白名单维度为手机号，配置含指定手机号\n3. 该手机号对应的用户账号',
    '1. 以白名单手机号用户调用授信进件提交\n2. 以同用户调用借款提交\n3. 查询partner_agreement_instance表',
    '1. 授信进件和借款提交均走新模式\n2. partner_agreement_instance表有协议记录\n3. 流程中不触发ags协议服务',
    '中', '否', '携程,灰度,白名单,手机号')

add('JYSG-149011', NODE,
    '携程灰度模式-身份证号维度白名单控制',
    '验证以身份证号为白名单维度的灰度控制，用户走新模式',
    '1. Apollo配置：partner_agreement_config开关关闭\n2. 灰度白名单维度为身份证号，配置含指定身份证号',
    '1. 以白名单身份证号用户调用借款提交\n2. 查询partner_agreement_instance表',
    '1. 借款提交走新模式\n2. partner_agreement_instance表有协议记录',
    '低', '否', '携程,灰度,白名单,身份证号')

add('JYSG-149011', NODE,
    '携程灰度模式-order_no维度白名单控制',
    '验证以order_no为白名单维度的灰度控制，订单走新模式',
    '1. Apollo配置：partner_agreement_config开关关闭\n2. 灰度白名单维度为order_no，配置含指定order_no',
    '1. 以白名单order_no调用借款提交\n2. 查询partner_agreement_instance表',
    '1. 该order_no走新模式\n2. partner_agreement_instance表有协议记录',
    '低', '否', '携程,灰度,白名单,order_no')

# ===== WRITE EXCEL =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row_idx, case in enumerate(cases, 2):
    values = [
        case['story'], case['node'], case['name'], case['summary'],
        case['precondition'], case['steps'], case['expected'],
        '', case['priority'], case['author'], case['type'],
        case['regression'], case['app'], case['tags']
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = cell_font
        cell.alignment = wrap
        cell.border = thin_border
        if col_idx in (1, 9, 10, 11, 12, 13):
            cell.alignment = center_align

widths = [22, 30, 38, 46, 50, 56, 56, 18, 8, 8, 10, 12, 10, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.row_dimensions[1].height = 28
for r in range(2, len(cases) + 2):
    ws.row_dimensions[r].height = 72

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:N{len(cases) + 1}'

output_path = r'D:\测试用例\20260611测试用例\携程协议_优化版.xlsx'
wb.save(output_path)

print(f'Saved: {output_path}')
print(f'Total cases: {len(cases)}')
high = sum(1 for c in cases if c['priority'] == '高')
mid = sum(1 for c in cases if c['priority'] == '中')
low = sum(1 for c in cases if c['priority'] == '低')
reg = sum(1 for c in cases if c['regression'] == '是')
print(f'Priority: 高={high}, 中={mid}, 低={low}')
print(f'Regression=是: {reg}')
print(f'Coverage: 新模式授信={sum(1 for c in cases if "授信" in c["name"])}, 新模式借款={sum(1 for c in cases if "借款" in c["name"])}, 签署={sum(1 for c in cases if "签署" in c["name"])}, 高登={sum(1 for c in cases if "高登" in c["name"])}, 灰度={sum(1 for c in cases if "灰度" in c["name"])}')
